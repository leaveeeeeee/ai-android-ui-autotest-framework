"""从结构化文本生成 pytest 用例与 AI 提示词。"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from framework.generator.models import TextCaseSpec
from framework.generator.renderer import render_ai_prompt, render_test_case


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="根据 Excel 或 CSV 测试文本生成 pytest 用例。")
    parser.add_argument("input_file", help="输入文件路径，支持 .xlsx 和 .csv。")
    parser.add_argument(
        "--output-dir",
        default="tests/generated",
        help="生成的 pytest 文件输出目录。",
    )
    parser.add_argument(
        "--prompt-dir",
        default="artifacts/report_data/ai-prompts",
        help="当用例信息不完整时，AI 提示词输出目录。",
    )
    return parser.parse_args()


def load_rows(input_path: Path) -> list[dict[str, str]]:
    """读取 CSV 或 XLSX 中的测试文本。"""
    if input_path.suffix.lower() == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as file:
            return list(csv.DictReader(file))

    if input_path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - 运行环境可能未安装 openpyxl
            raise SystemExit("读取 .xlsx 需要 openpyxl，请先安装项目依赖。") from exc

        workbook = load_workbook(input_path)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [
            {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            for row in rows[1:]
        ]

    raise SystemExit("目前只支持 .xlsx 和 .csv 两种输入格式。")


def build_manifest_path(input_path: Path, output_dir: Path) -> Path:
    """为当前输入文件生成稳定的 manifest 路径。"""
    output_root = output_dir.resolve()
    source_key = hashlib.sha1(str(input_path.resolve()).encode("utf-8")).hexdigest()[:12]
    safe_name = "".join(char if char.isalnum() else "_" for char in input_path.stem).strip("_")
    manifest_dir = output_root / ".manifest"
    manifest_dir.mkdir(parents=True, exist_ok=True)
    return manifest_dir / f"{safe_name or 'generated'}_{source_key}.json"


def load_manifest(path: Path) -> dict[str, list[str]]:
    """读取上一次生成记录。"""
    if not path.exists():
        return {"generated_files": [], "prompt_files": []}
    return json.loads(path.read_text(encoding="utf-8"))


def cleanup_stale_outputs(
    previous_manifest: dict[str, list[str]],
    *,
    current_generated: set[str],
    current_prompts: set[str],
    output_dir: Path,
    prompt_dir: Path,
) -> tuple[int, int]:
    """清理当前输入源历史生成、但本次已经不存在的文件。"""
    removed_generated = 0
    removed_prompts = 0
    targets = (
        ("generated_files", current_generated, output_dir.resolve()),
        ("prompt_files", current_prompts, prompt_dir.resolve()),
    )
    for key, current_paths, allowed_root in targets:
        for previous_path in previous_manifest.get(key, []):
            target = Path(previous_path)
            if previous_path in current_paths or not target.exists():
                continue
            resolved = target.resolve()
            if allowed_root not in (resolved, *resolved.parents):
                continue
            target.unlink()
            if key == "generated_files":
                removed_generated += 1
            else:
                removed_prompts += 1
    return removed_generated, removed_prompts


def write_manifest(
    path: Path,
    *,
    input_path: Path,
    generated_files: set[str],
    prompt_files: set[str],
) -> None:
    """保存当前输入源对应的生成记录。"""
    payload = {
        "input_file": str(input_path.resolve()),
        "generated_files": sorted(generated_files),
        "prompt_files": sorted(prompt_files),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    """执行文本转用例主流程。"""
    args = parse_args()
    input_path = Path(args.input_file)
    output_dir = Path(args.output_dir)
    prompt_dir = Path(args.prompt_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    prompt_dir.mkdir(parents=True, exist_ok=True)

    rows = load_rows(input_path)
    manifest_path = build_manifest_path(input_path, output_dir)
    previous_manifest = load_manifest(manifest_path)
    generated = 0
    prompts = 0
    generated_files: set[str] = set()
    prompt_files: set[str] = set()
    for row in rows:
        spec = TextCaseSpec.from_mapping(row)
        target_file = output_dir / spec.file_name
        target_file.write_text(render_test_case(spec), encoding="utf-8")
        generated_files.add(str(target_file.resolve()))
        generated += 1

        if not spec.has_executable_calls:
            prompt_file = prompt_dir / f"{spec.case_id}.md"
            prompt_file.write_text(render_ai_prompt(spec), encoding="utf-8")
            prompt_files.add(str(prompt_file.resolve()))
            prompts += 1

    removed_generated, removed_prompts = cleanup_stale_outputs(
        previous_manifest,
        current_generated=generated_files,
        current_prompts=prompt_files,
        output_dir=output_dir,
        prompt_dir=prompt_dir,
    )
    write_manifest(
        manifest_path,
        input_path=input_path,
        generated_files=generated_files,
        prompt_files=prompt_files,
    )

    print(f"已生成 pytest 文件数：{generated}")
    print(f"已生成 AI 提示词数：{prompts}")
    print(f"已清理过期 pytest 文件数：{removed_generated}")
    print(f"已清理过期 AI 提示词数：{removed_prompts}")
    print(f"用例输出目录：{output_dir.resolve()}")
    print(f"提示词输出目录：{prompt_dir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
